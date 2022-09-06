import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
import numpy as np
from sklearn import ensemble, model_selection
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.model_selection import RandomizedSearchCV#ランダムサーチ
from sklearn.ensemble import RandomForestClassifier as RFC#ランダムフォレスト(クラス分類用)
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix,precision_score,recall_score,f1_score
from sklearn.metrics import r2_score#決定係数求める用
import joblib#モデルの保存、ロード用
from sklearn.model_selection import learning_curve#学習曲線
import matplotlib.pyplot as plt
import glob
import random
import os
import openpyxl

for l in range(1):
    if l == 0:
        files = glob.glob('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\now_all\\*')
    elif l == 1:
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
    else :
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
            
    df_list_test = []
    df_list_train = []
    files = glob.glob('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\now_all\\*')
    test = random.sample(files,10)
    for file in files:
        if file in test:
            t=0
            path, ext = os.path.splitext(file)
            sub_df = pd.read_csv(file)
            df_list_test.append(sub_df)
            print(sub_df)
        else :
            t=0
            path, ext = os.path.splitext(file)
            sub_df = pd.read_csv(file)
            df_list_train.append(sub_df)
            print(sub_df)

    df_train2 = pd.concat(df_list_train)
    df_test2 = pd.concat(df_list_test)


    print(1)
    df_train2.columns

    print(df_train2)
    print(df_test2)

    criterion_array=[]
    max_depth_array=[]
    max_features_array=[]
    min_samples_split_array=[]
    min_samples_leaf_array=[]
    n_estimators_array =[]
    n_jobs_array=[]

    poms=["workload"]
    #グリッドサーチ、分類のバランスを考慮、ラベルエンコーディングしたとき
    #https://qiita.com/kazuki_hayakawa/items/6d7a4597829f54ebdb83
    for i in poms:
        #テストデータと訓練データを読み取り
        #df_train2=pd.read_csv("train_total_labelencording_15T_person_pomschange.csv",index_col=0)
        #df_test2=pd.read_csv("test_total_labelencording_15T_person_pomschange.csv",index_col=0)
        
        #テストデータと訓練データを読み取り
        train_dataset=df_train2
        test_dataset=df_test2

        #使わない列を削除
        #df_train2=train_dataset.drop(["   nwhel,steer ","   _trim,_elev "], axis=1)
        #df_test2=test_dataset.drop(["   nwhel,steer ","   _trim,_elev "], axis=1)
        
        
        #特徴量
        name=df_train2.columns
        
        #訓練データとテストデータに分ける
        x_train = df_train2.drop(i, axis=1).values
        y_train= df_train2[i].values
        x_test=df_test2.drop(i, axis=1).values
        y_test= df_test2[i].values
        
        #データを整形する

        Xtrain2 = np.array(x_train)
        Xtest2 = np.array(x_test)
        y_train2 = np.array(y_train)
        y_test2 = np.array(y_test)
        
        search_params = {
        #'criterion': ['gini', 'entropy'],
        #'n_estimators'      : [10,20,50,100],#決定木の数
        #'max_features'      : ['sqrt', 'log2'],#個々の決定木に使用する特徴量の数
        #'n_jobs'            : [1],
        #'min_samples_split' : [3, 5, 10, 20, 30, 40, 50, 100],
        #'min_samples_leaf': [5,50],
        #'max_depth'         : [5,10,20,50,100]#決定木最大の深さ

        'criterion': ['entropy'],
        'n_estimators'      : [40,50,60],#決定木の数
        'max_features'      : ['sqrt'],#個々の決定木に使用する特徴量の数
        #'n_jobs'            : [1],
        #'min_samples_split' : [3, 5, 10, 20, 30, 40, 50, 100],
        'min_samples_leaf': [5,50],
        'max_depth'         : [5,10,15,18]#決定木最大の深さ
        }
        
        RFC_grid = GridSearchCV(estimator=RFC(random_state=0,class_weight='balanced'),param_grid=search_params, scoring='accuracy', cv=5,verbose=True,n_jobs=-1) #グリッドサーチ・ランダムフォレスト
        
        RFC_grid.fit(Xtrain2, y_train2)
        best_model = RFC_grid.best_estimator_
        best_model_params=RFC_grid.best_params_
        
        criterion_ = best_model_params["criterion"]
        max_depth_ = best_model_params["max_depth"]
        max_features_=best_model_params["max_features"]
        #min_samples_split_ = best_model_params["min_samples_split"]
        #min_samples_leaf_=best_model_params["min_samples_leaf"]
        n_estimators_ = best_model_params["n_estimators"]
        #n_jobs_=best_model_params["n_jobs"]
        
        #パラメータを配列にいれる
        criterion_array.append(criterion_)
        max_depth_array.append(max_depth_)
        max_features_array.append(max_features_)
        #min_samples_split_array.append(min_samples_split_)
        #min_samples_leaf_array.append(min_samples_leaf_)
        n_estimators_array .append(n_estimators_)
        #n_jobs_array.append(n_jobs_)
        
        #モデルの構築
        model2 = RFC(class_weight='balanced',criterion=criterion_, max_depth = max_depth_, max_features=max_features_ , n_estimators=n_estimators_,random_state=0,verbose=1)
        #model2 = RFC(class_weight='balanced',criterion=criterion_, max_depth = max_depth_, max_features=max_features_,min_samples_split= min_samples_split_,n_estimators=n_estimators_,n_jobs=n_jobs_,random_state=0,verbose=1)

        # モデルのコンパイル
        #model_.compile(loss='mse',
        #              optimizer=tf.keras.optimizers.Adam(learning_rate=lr_),
        #              metrics=['mae', 'mse'])

        # モデルの学習
        history_ = model2.fit(Xtrain2, y_train2)
        
        # モデルを保存する(https://localab.jp/blog/save-and-load-machine-learning-models-in-python-with-scikit-learn/)
        #filename = 'RDF_Classification_grid'+i+'finalized_model.sav'
        #joblib.dump(history_, open(filename, 'wb'))
        
        pred_train = history_.predict(Xtrain2)
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_train = model2.predict(Xtrain2)
        print("グリッドサーチ・ランダムフォレスト")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(y_train2, pred_train,output_dict = True)
        print( train_class )

        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(y_train2, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(y_train2, pred_train) )
        
        print(i+'モデルのグリッドサーチ・ランダムフォレストモデルにおける n_estimators   :  %d'  %RFC_grid.best_estimator_.n_estimators)
        
        print("グリッドサーチ・ランダムフォレスト")
        print(RFC_grid.best_estimator_)#最も性能がよかったランダムフォレストのインスタンス
        print('Best params: {}'.format(RFC_grid.best_params_)) 
        print('Best Score: {}'.format(RFC_grid.best_score_))#'criterion': 'entropy'の場合は小さいほうが良い、gridの場合は大きいほうがよい

        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_test = model2.predict(Xtest2)
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(y_test2, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(y_test2, pred_test)
        print( test_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(y_test2, pred_test) )
        
        feature = model2.feature_importances_
        # 特徴量の名前ラベルを取得
        label = df_train2.columns[0:]
        # 特徴量の重要度順（降順）に並べて表示
        indices = np.argsort(feature)[::-1]
        for i in range(len(feature)):
            print(str(i + 1) + "   " +
                str(label[indices[i]]) + "   " + str(feature[indices[i]]))
            
        # 実際の値と予測値の比較グラフ
        plt.subplot(121, facecolor='white')
        #plt_label = [i for i in range(1, 32)]
        plt.plot(y_test2, color='blue')
        plt.plot(pred_test, color='red')
        # 特徴量の重要度の棒グラフ
        plt.subplot(122, facecolor='white')
        plt.title('特徴量の重要度')
        plt.bar(
            range(
                len(feature)),
            feature[indices],
            color='blue',
            align='center')
        plt.xticks(range(len(feature)), label[indices], rotation=45)
        plt.xlim([-1, len(feature)])
        plt.tight_layout()
        # グラフの表示
        plt.show()
        print(y_test2)
        print(pred_test)
        y_test2 = pd.DataFrame(y_test2).T
        pred_test = pd.DataFrame(pred_test).T
        y_test2.to_csv('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\output\\test.csv')
        pred_test.to_csv('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\output\\pred.csv')

        report_df = pd.DataFrame(train_class).T
        report_df.to_csv('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\output\\rdf_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\output\\rdf_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(j+1,k+1,value = train_conf[j][k])
                ws.cell(j+7,k+1,value = test_conf[j][k])
        
        wb.save('C:\\Users\\rikua\\Documents\\ml_run_fdToWl\\ml_all\\output\\rdf'+ str(l) + '.xlsx')